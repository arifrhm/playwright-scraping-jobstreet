from openpyxl import load_workbook
import json

# Membaca file Excel yang sudah ada
excel_file_path = "Usulan RKAP (Anggaran dan Target).xlsx"
workbook = load_workbook(excel_file_path, data_only=True)  # `data_only=True` untuk mendapatkan nilai bukan rumus

# Dictionary untuk menyimpan hasil mapping
excel_data = {}

# Iterasi setiap sheet di workbook
for sheet in workbook.sheetnames:
    worksheet = workbook[sheet]
    sheet_data = {}

    # Iterasi setiap baris dan kolom di worksheet
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                # Menggunakan `cell.coordinate` untuk mendapatkan alamat sel (misal, A1, B2, dll.)
                sheet_data[cell.coordinate] = cell.value

    # Menyimpan data setiap sheet ke dalam dictionary
    excel_data[sheet] = sheet_data

# Menyimpan data ke file JSON
output_file = "output_mapping.json"
with open(output_file, 'w') as json_file:
    json.dump(excel_data, json_file, indent=4)

# Menampilkan hasil mapping
print(json.dumps(excel_data, indent=4))

# Menutup workbook
workbook.close()
