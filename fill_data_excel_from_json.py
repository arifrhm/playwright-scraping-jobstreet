import json
from openpyxl import load_workbook

# Membaca file JSON Proposal Riset (data yang akan diisi ke Excel)
proposal_file_path = "proposal-riset-all.json"
with open(proposal_file_path, 'r') as proposal_file:
    proposal_data = json.load(proposal_file)

# Membaca file JSON Output Mapping (mapping sel di Excel)
mapping_file_path = "output_mapping.json"
with open(mapping_file_path, 'r') as mapping_file:
    mapping_data = json.load(mapping_file)

# Membaca file Excel salinan yang akan diisi
excel_file_path = "Usulan RKAP (Anggaran dan Target) copy.xlsx"
workbook = load_workbook(excel_file_path)

# Iterasi setiap sheet yang terdapat dalam mapping
for sheet_name, cells in mapping_data.items():
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        # Mengisi setiap cell berdasarkan informasi dari JSON mapping
        for cell_address, json_key in cells.items():
            # Key di sini mengacu pada kunci yang ada di JSON proposal riset
            value = proposal_data.get(json_key)

            # Hanya isi jika value ditemukan di JSON
            if value is not None:
                worksheet[cell_address] = value

# Menyimpan workbook yang telah diperbarui
workbook.save(excel_file_path)
workbook.close()

print("Data dari JSON berhasil diisikan ke dalam file Excel.")
